import os
import pickle
import numpy as np
import cv2
import face_recognition
import cvzone
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import storage
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import sys
import time

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred, {
    'databaseURL': "https://entrysystem-276b3-default-rtdb.firebaseio.com/",
    'storageBucket': "entrysystem-276b3.appspot.com"
})

bucket = storage.bucket()

cap = cv2.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

imgBackground = cv2.imread('Resources/background.png')

# Importing the mode images into a list
folderModePath = 'Resources/Modes'
modePathList = os.listdir(folderModePath)
imgModeList = []
for path in modePathList:
    imgModeList.append(cv2.imread(os.path.join(folderModePath, path)))

# Load the encoding file
print("Loading Encode File ...")
file = open('EncodeFile.p', 'rb')
encodeListKnownWithIds = pickle.load(file)
file.close()
encodeListKnown, employeeIds = encodeListKnownWithIds
print("Encode File Loaded")

modeType = 0
counter = 0
id = -1
imgEmployee = []

# Boolean variable for Face Recognition status
face_recognition_status = False

# Entry form variables
entry_form_opened = False
entry_form_start_time = None
entry_form_time_limit = 20  # Set the time limit for the entry form in seconds

# Set the maximum time limit for the entry form appearance
max_entry_form_time = 15  # seconds

# Record the start time at the beginning
start_time = datetime.now()

while True:
    # Calculate elapsed time
    elapsed_time = (datetime.now() - start_time).total_seconds()

    # Open entry form when elapsed time exceeds 15 seconds
    if elapsed_time > max_entry_form_time and not entry_form_opened:
        entry_form_opened = True

        def save_data():
            # Get the user input from the entry fields
            name = name_entry.get()
            last_name = last_name_entry.get()
            contact = contact_entry.get()
            post = post_entry.get()

            # Check if all fields are filled
            if not all([name, last_name, contact, post]):
                messagebox.showwarning("Incomplete Form", "Please fill in all fields.")
                return

            try:
                wb = Workbook()
                try:
                    wb = load_workbook("Visitors.xlsx")
                except FileNotFoundError:
                    pass

                sheet = wb.active  # Get the active worksheet

                # Write the column headers if the file is newly created
                if not sheet["A1"].value:
                    sheet["A1"] = "First Name"
                    sheet["B1"] = "Last Name"
                    sheet["C1"] = "Contact"
                    sheet["D1"] = "Post"
                    sheet["E1"] = "Timestamp"

                # Get the current timestamp and date
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Write the user input and timestamp to the next available row in the Excel sheet
                row_num = sheet.max_row + 1
                sheet.cell(row=row_num, column=1).value = name
                sheet.cell(row=row_num, column=2).value = last_name
                sheet.cell(row=row_num, column=3).value = contact
                sheet.cell(row=row_num, column=4).value = post
                sheet.cell(row=row_num, column=5).value = timestamp

                # Save the Excel file
                wb.save("Visitors.xlsx")

                # Inform the user that the data has been submitted
                messagebox.showinfo("Submission Successful", "Visitor information has been successfully submitted.")

                # Close the program
                window.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")

        # Create the GUI window
        window = tk.Tk()
        window.title("Visitor Information Form")
        window.geometry("400x250")  # Set window size

        # Create a frame for better organization
        frame = ttk.Frame(window, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Create labels and entry fields for the form
        name_label = ttk.Label(frame, text="First Name:")
        name_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        name_entry = ttk.Entry(frame)
        name_entry.grid(row=0, column=1, padx=10, pady=5)

        last_name_label = ttk.Label(frame, text="Last Name:")
        last_name_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        last_name_entry = ttk.Entry(frame)
        last_name_entry.grid(row=1, column=1, padx=10, pady=5)

        contact_label = ttk.Label(frame, text="Contact:")
        contact_label.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        contact_entry = ttk.Entry(frame)
        contact_entry.grid(row=2, column=1, padx=10, pady=5)

        post_label = ttk.Label(frame, text="Post:")
        post_label.grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
        post_entry = ttk.Entry(frame)
        post_entry.grid(row=3, column=1, padx=10, pady=5)

        # Create a button to submit the form
        submit_button = ttk.Button(frame, text="Submit", command=save_data)
        submit_button.grid(row=4, column=0, columnspan=2, pady=10)

        # Start the GUI event loop
        window.mainloop()
        sys.exit()

    success, img = cap.read()

    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    faceCurFrame = face_recognition.face_locations(imgS)
    encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

    imgBackground[162:162 + 480, 55:55 + 640] = img
    imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]

    if faceCurFrame:
        for encodeFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                bbox = 55 + x1, 162 + y1, x2 - x1, y2 - y1
                imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)
                id = employeeIds[matchIndex]
                if counter == 0:
                    cvzone.putTextRect(imgBackground, "Loading", (275, 400))
                    cv2.imshow("Face Attendance", imgBackground)
                    cv2.waitKey(1)
                    counter = 1
                    modeType = 1
                    entry_form_start_time = datetime.now()

                # Successful entry, set attendance status to True
                attendance_status = True

        if counter != 0:
            if counter == 1:
                # Get the Data
                employeeInfo = db.reference(f'Employee/{id}').get()
                print(employeeInfo)

                blob = bucket.get_blob(f'Images/{id}.png')
                array = np.frombuffer(blob.download_as_string(), np.uint8)
                imgEmployee = cv2.imdecode(array, cv2.COLOR_BGRA2BGR)

                datetimeObject = datetime.strptime(employeeInfo['last_attendance_time'], "%Y-%m-%d %H:%M:%S")
                secondsElapsed = (datetime.now() - datetimeObject).total_seconds()
                print(secondsElapsed)
                if secondsElapsed > 30:
                    ref = db.reference(f'Employee/{id}')
                    employeeInfo['total_attendance'] += 1
                    ref.child('total_attendance').set(employeeInfo['total_attendance'])
                    ref.child('last_attendance_time').set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    ref.child('attendance_status').set(attendance_status)  # Set attendance status to True
                    # Set face_recognition_status to True after successful upload
                    face_recognition_status = True
                else:
                    modeType = 3
                    counter = 0
                    imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]

            if modeType != 3:
                if 10 < counter < 20:
                    modeType = 2

                imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]

                if counter <= 10:
                    cv2.putText(imgBackground, str(employeeInfo['total_attendance']), (861, 125),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
                    cv2.putText(imgBackground, str(employeeInfo['major']), (1006, 550),
                                cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(imgBackground, str(id), (1006, 493),
                                cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(imgBackground, str(employeeInfo['standing']), (910, 625),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)
                    cv2.putText(imgBackground, str(employeeInfo['year']), (1025, 625),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)
                    cv2.putText(imgBackground, str(employeeInfo['starting_year']), (1125, 625),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)

                    (w, h), _ = cv2.getTextSize(employeeInfo['name'], cv2.FONT_HERSHEY_COMPLEX, 1, 1)
                    offset = (414 - w) // 2
                    cv2.putText(imgBackground, str(employeeInfo['name']), (808 + offset, 445),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (50, 50, 50), 1)

                    imgBackground[175:175 + 216, 909:909 + 216] = imgEmployee
                    

                counter += 1

                if counter >= 20:
                    counter = 0
                    modeType = 0
                    employeeInfo = []
                    imgEmployee = []
                    imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]
    else:
        modeType = 0
        counter = 0
       
    # Check if face recognition is successful
    if face_recognition_status:
        print("Face Recognition Successful!")
        # Your further actions for successful face recognition can be added here
        # Reset variables for the next recognition cycle
        face_recognition_status = False
        time.sleep(3)  
        sys.exit(0)
        

    # Check if 20 seconds have passed without face recognition
    if counter == 0 and entry_form_start_time:
        seconds_elapsed_entry_form = (datetime.now() - entry_form_start_time).total_seconds()
        if seconds_elapsed_entry_form > entry_form_time_limit and not entry_form_opened:
            entry_form_opened = True
            print("Opening Entry Form")

            def save_data():
                # Get the user input from the entry fields
                name = name_entry.get()
                last_name = last_name_entry.get()
                contact = contact_entry.get()
                post = post_entry.get()

                # Check if all fields are filled
                if not all([name, last_name, contact, post]):
                    messagebox.showwarning("Incomplete Form", "Please fill in all fields.")
                    return

                try:
                    wb = Workbook()
                    try:
                        wb = load_workbook("Visitors.xlsx")
                    except FileNotFoundError:
                        pass

                    sheet = wb.active  # Get the active worksheet

                    # Write the column headers if the file is newly created
                    if not sheet["A1"].value:
                        sheet["A1"] = "First Name"
                        sheet["B1"] = "Last Name"
                        sheet["C1"] = "Contact"
                        sheet["D1"] = "Post"
                        sheet["E1"] = "Timestamp"

                    # Get the current timestamp and date
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    # Write the user input and timestamp to the next available row in the Excel sheet
                    row_num = sheet.max_row + 1
                    sheet.cell(row=row_num, column=1).value = name
                    sheet.cell(row=row_num, column=2).value = last_name
                    sheet.cell(row=row_num, column=3).value = contact
                    sheet.cell(row=row_num, column=4).value = post
                    sheet.cell(row=row_num, column=5).value = timestamp

                    # Save the Excel file
                    wb.save("Visitors.xlsx")

                    # Inform the user that the data has been submitted
                    messagebox.showinfo("Submission Successful", "Visitor information has been successfully submitted.")

                    # Close the program
                    window.destroy()

                except Exception as e:
                    messagebox.showerror("Error", f"An error occurred: {str(e)}")

            # Create the GUI window
            window = tk.Tk()
            window.title("Visitor Information Form")
            window.geometry("400x250")  # Set window size

            # Create a frame for better organization
            frame = ttk.Frame(window, padding="10")
            frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

            # Create labels and entry fields for the form
            name_label = ttk.Label(frame, text="First Name:")
            name_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
            name_entry = ttk.Entry(frame)
            name_entry.grid(row=0, column=1, padx=10, pady=5)

            last_name_label = ttk.Label(frame, text="Last Name:")
            last_name_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
            last_name_entry = ttk.Entry(frame)
            last_name_entry.grid(row=1, column=1, padx=10, pady=5)

            contact_label = ttk.Label(frame, text="Contact:")
            contact_label.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
            contact_entry = ttk.Entry(frame)
            contact_entry.grid(row=2, column=1, padx=10, pady=5)

            post_label = ttk.Label(frame, text="Post:")
            post_label.grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
            post_entry = ttk.Entry(frame)
            post_entry.grid(row=3, column=1, padx=10, pady=5)

            # Create a button to submit the form
            submit_button = ttk.Button(frame, text="Submit", command=save_data)
            submit_button.grid(row=4, column=0, columnspan=2, pady=10)

            # Start the GUI event loop
            window.mainloop()
            time.sleep(3)
            sys.exit()

    cv2.imshow("Face Attendance", imgBackground)
    cv2.waitKey(1)
